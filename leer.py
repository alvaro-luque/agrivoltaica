#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Nov 16 09:52:03 2021

@author: Álvaro Luque Sánchez
"""

import numpy as np
import openpyxl as pxl
import matplotlib.pyplot as plt
import mod1
import time


#esto nos quita la notación científica de numpy
np.set_printoptions(suppress=True)
#
# para medir el tiempo
# start=time.time()
# print("--- %s seconds ---" % (time.time() - start))
#
calc=pxl.load_workbook('ejercicio.xlsx')

#leemos de la hoja 'datos'; cargamos los datos en hoja 1 (h1)

h1=calc['AgriPV']

#vamos a guardar todos los datos en un array (6x12) de numpy
#6 columnas de datos, 12 filas (una por mes); lo inicializamos con zeros

datos=np.zeros((6,12))

for i in range(6):
    for j in range(12):
         #las siguientes líneas comprobarían que el dato es un número, 
         #pero no es necesario ya que vamos a ajustar el bucle a la hoja de cálculo
         #if h1.cell(j+1,i+1).data_type != 'n':
         #  continue
         #para ajustar los índices, simplemente escribir las coordenadas (fila,columna)
         #de la primera celda que queramos leer
        datos[i,j]=h1.cell(j+5,i+2).value
lat=h1['B1'].value
long=h1['E1'].value
albedo=h1['B3'].value

#ahora vamos a montar las placas que se indican en la hoja AgriPV (el suelo no es necesario de momento)
#para ello necesitamos extraer las coordenadas de los 3 puntos que definen a cada una

puntos=[]


for i in range(9):
    puntos.append((h1['B%.2s' % str(22+i)].value, h1['C%.2s' % str(22+i)].value, h1['D%.2s' % str(22+i)].value))
    puntos.append((h1['E%.2s' % str(22+i)].value, h1['F%.2s' % str(22+i)].value, h1['G%.2s' % str(22+i)].value))
    puntos.append((h1['H%.2s' % str(22+i)].value, h1['I%.2s' % str(22+i)].value, h1['J%.2s' % str(22+i)].value))

placa1=puntos[0:3]
placa2=puntos[3:6]
placa3=puntos[6:9]
placa4=puntos[9:12]
placa5=puntos[12:15]
placa6=puntos[15:18]
placa7=puntos[18:21]
placa8=puntos[21:24]
placa9=puntos[24:27]


# GMSH crea rectángulos dando su "origen", es decir, su esquina inferior izquierda. Creamos estas variables
# de paso para que sean más accesibles estos puntos luego
# Esta esquina es simplemente la que tiene las "coordenadas más pequeñas".
origen1=min(placa1)
origen2=min(placa2)

placas=[placa1, placa2, placa3, placa4, placa5, placa6, placa7, placa8, placa9]
placass=[]
i=0
for placa in placas:
    placas[i]=mod1.Placa(placa[0],placa[1],placa[2])    
    placass.append(placas[i])
    i+=1
    
placas=placass

del placass
    
horas=np.arange(-12,12,1)

#dia 15 de cada mes:
dias=np.array([mod1.diajuliano(15, i+1, 2022) for i in range(12)])

#aquí calculamos TODOS los vectores solares del día 15 de los 12 meses del año
#vamos a separar los datos por meses con reshape(12, 24, 3) 24 entradas de vectorsolar por cada día
#de cada mes, 3 componentes. solar[i][j][k] es la componente k del vector solar a la hora j del 
#mes i (ajuste orígenes índices)

solar=np.array([mod1.vectorsolar(dia, hora, lat) for dia in dias for hora in horas])\
                                            .reshape(dias.shape[0],horas.shape[0],3)

#tenemos 24 horas al dia (de -12 a +11 pasando por el 0), por 1 día/mes, 24 entradas de solar por mes
#para el día escogido (el 15) del mes j, cogemos solar[24(j-1):24j] 
#hacer filtrado de z positivo DESPUÉS de seleccionar el mes para ahorrar cálculos y mantener la fórmula


#ajustar a las dimensiones del terreno y número de puntos
mallax=np.linspace(0,60,60)
mallay=np.linspace(0,40,40)
malla=np.array([[x,y, 0] for x in mallax for y in mallay])
factores=np.zeros((dias.shape[0], horas.shape[0], malla.shape[0])) #un svf por punto 
#un svf por hora y día para tener en cuenta la rotación de las placas

#lista que luego transformamos a array una vez proyectada
esfera=[[i/18, j/18, 0] for i in range(-18,19) for j in range(-18,19)]

for k in range(len(esfera)):
    if esfera[k][0]**2+esfera[k][1]**2<1:
        esfera[k][2]=np.sqrt(1-esfera[k][0]**2-esfera[k][1]**2)
    else:
       continue 

esfera=np.array(esfera)
#dejamos los puntos que se queden fuera de la semiesfera con z=0, y luego los omitimos con una condición

tiempo=0


for dia in range(solar.shape[0]):
        
    for hora in range(solar.shape[1]):
        
        if solar[dia][hora][2]>0:
            for placa in placas:
                placa.crear_rotada(solar[dia][hora]) #normal=vector solar en cada momento
            
            j=0   
            for punto in malla:    
                start=time.time()
                cero=0
                uno=0    
                for rayo in esfera:
                    if rayo[2]==0:
                        continue
                    
                    else:
                        for i in range(len(placas)):
                            sec=placas[i].interseca(punto,rayo) 
                            print(sec)
                            
                            if sec==0 and i+1==len(placas): 
#hay que sumar uno ya que es un índice y un objeto de tamaño 2 tiene índices 0 y 1.
                                cero +=1
                            elif sec==1:
                                uno+=1
                                break #si interseca con una placa, hemos terminado con ese rayo
                
                factores[dia, hora, j]=cero/(cero+uno)
                print(factores[dia,hora,j])
                if j+1==malla.shape[0]:
                    break
                else:
                    j+=1    
                tiempo += time.time()-start
                print('Tiempo para el punto (%.2f, %.2f): %.4f' % (punto[0],punto[1], time.time()-start)) 
        else:
            continue
    for placa in placas:
        placa.reiniciar()
        
#factores tendrá muchos arrays vacíos en él. Podríamos borrarlos pero eso nos da problemas luego para saber
#a qué día y hora corresponde cada uno. 
#Para representar un mapa de calor rápido:
# facc=factores[dia][hora].reshape(mallax.size,mallay.size).transpose()
# fig=plt.figure()
# ax=fig.add_subplot()
# mapa=ax.imshow(facc)
# plt.colorbar(mapa)

#gestión de datos de radiación
tmy=pxl.load_workbook('tmy.xlsx')

hojadatos=tmy['datos']

#en tmy tenemos un dato horario por cada día del mes, extraemos solo los días 15 de estos
#los datos del 1 de enero comienzan en la fila 18. A 24 datos por día, el día 2 empieza en la fila 42
#el día juliano k irá desde el 24k-6 hasta el 24k+17, k=1,...,365
#en nuestro caso, tenemos la variable dias de forma que dias[j-1] con j=1,...,12 es el día 15 del mes j
#luego podemos coger los días que tenemos haciendo k=dias[j-1]

#de la hoja de datos nos interesan las columnas D y F (G y Gd respectivamente)

G=[]
Gd=[]
start=time.time()
#queremos extraer los datos del día 15 de cada mes, introducimos una lista auxiliar
#que almacenará los datos de cada día, luego los volcamos en la correspondiente G o Gd
for dia in dias:
    listag=[]
    listagd=[]
    for col in hojadatos.iter_cols(min_row=24*dia-6, max_row=24*dia+17, min_col=4, max_col=6):
        for cell in col:
            if cell.column==4:
                listag.append(cell.value)
                
            elif cell.column==5:
                continue
            
            elif cell.column==6:
                listagd.append(cell.value)
    G.append(listag)
    Gd.append(listagd)       
    
G=np.array(G)
Gd=np.array(Gd)
Gb=G-Gd
print('Tiempo: %s segundos' % str(time.time()-start))

rad=np.zeros(factores) #alberga los datos de radiacion finales
#ya tenemos la G y la Gd guardadas en arrays columna (reshape) para facilitar 
#los cálculos ahora

for dia in range(12):
    for hora in range(24):
        rad[i][j]=Gb[i][j]+Gd[i][j]*factores[i][j]

#así, tenemos en la variable rad los datos de radiación que necesitamos
#de nuevo, tiene muchas componentes nulas, hay que sistematizar el filtrado
#y organización de los datos


calc.close()
#el método close() debe usarse cuando se terminen de escribir los datos
#en Excel.

#vamos a realizar el modelo gráfico usando la API de Python de Gmesh.
#En concreto, usaremos el kernel gráfico de OpenCASCADE que nos permite hacer geometría
#sólida "más fácilmente" que el Gmesh nativo.

# import gmsh

# gmsh.initialize()

# gmsh.model.add('huerto')

# gmsh.model.occ.addRectangle(0,0,0,100,100) #este es el suelo

# gmsh.model.occ.addPoint(placa1.lado2.midpoint.x, placa1.lado2.midpoint.y, placa1.lado2.midpoint.z, tag=45)
# gmsh.model.occ.addPoint(placa1.lado2.midpoint.x, placa1.lado2.midpoint.y, 0, tag=46)

# pata=gmsh.model.occ.addLine(45,46)
# gmsh.model.occ.translate(gmsh.model.occ.copy([(1,pata)]),0,placa1.largo,0)


# plc=gmsh.model.occ.addRectangle(origen1[0], origen1[1], origen1[2], placa1.largo, placa1.alto)
# #dado que todas las placas son iguales, solo vamos a crear una que moveremos con el método rotate/translate
# #de gmsh

# plc2=gmsh.model.occ.copy([(2,plc)])
   

# gmsh.model.occ.translate(gmsh.model.occ.copy([(1,pata)]),30,0,0)
# gmsh.model.occ.translate(gmsh.model.occ.copy([(1,pata)]),30,placa1.largo,0)    
# gmsh.model.occ.translate(plc2, 0,30,0) 

# gmsh.model.occ.synchronize()

# gmsh.fltk.run()
    
    
    
    
    
    
    
    
    
    
    
    
    
